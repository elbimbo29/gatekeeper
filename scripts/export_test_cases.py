"""
Gatekeeper QA Test Case Export Script
-------------------------------------

This script generates an Excel file (`test_cases.xlsx`) containing
a structured checklist of Gatekeeper app test cases.

Usage:
    1. Run from the project root:
        python scripts/export_test_cases.py

    2. The script will create `test_cases.xlsx` in the root folder.

    3. Open the file in Excel/Google Sheets to track QA results:
        - Fill in "Actual Result" after performing each test.
        - Mark "Status" as ✅ Pass or ❌ Fail.
        - Timestamp column auto-fills when Status is updated.
        - Tester Initials column lets multiple testers log who ran each case.
        - Conditional formatting is applied automatically:
            ✅ Pass → Green
            ❌ Fail → Red
        - Summary sheet shows totals, success percentage, and charts.

Purpose:
    - Provides a clean, professional QA log.
    - Easier to read and manage than Markdown tables.
    - Ensures all critical features (Signup, Login, Role Redirects,
      Remember Me, Session Expiry, Logout, Theme Toggle) are tested.
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, BarChart, Reference


def export_test_cases(filename="test_cases.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gatekeeper Test Cases"

    # Headers
    headers = [
        "Feature",
        "Test Case",
        "Expected Result",
        "Actual Result",
        "Status",
        "Timestamp",
        "Tester Initials",
    ]
    ws.append(headers)

    # Test cases
    test_cases = [
        ("Signup", "Create new user", "User added, success message", "", "", "", ""),
        (
            "Signup Dup",
            "Try existing username",
            "Error 'Username already exists'",
            "",
            "",
            "",
            "",
        ),
        (
            "Login Valid",
            "Correct username & password",
            "Redirect to dashboard",
            "",
            "",
            "",
            "",
        ),
        ("Login Invalid", "Wrong password", "Error message", "", "", "", ""),
        ("Login Unknown", "Non-existent username", "Error message", "", "", "", ""),
        ("Role Redirect", "Admin", "Goes to admin_dashboard.py", "", "", "", ""),
        ("Role Redirect", "Logs", "Goes to logs_dashboard.py", "", "", "", ""),
        ("Role Redirect", "User", "Goes to user_dashboard.py", "", "", "", ""),
        (
            "Unknown Role",
            "Invalid role in DB",
            "Error 'Unknown role. Logging out'",
            "",
            "",
            "",
            "",
        ),
        ("Remember Me", "Checked", "Session expiry = 7 days", "", "", "", ""),
        (
            "Session Expiry",
            "Observe sidebar countdown",
            "Shows remaining time + indicator",
            "",
            "",
            "",
            "",
        ),
        ("Logout", "Click logout", "Cookies cleared, back to login", "", "", "", ""),
        (
            "Theme Toggle",
            "Switch Dark/Light",
            "Colors update correctly",
            "",
            "",
            "",
            "",
        ),
    ]

    for case in test_cases:
        ws.append(case)

    # Apply conditional formatting to Status column (E)
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add(
        "E2:E100", CellIsRule(operator="equal", formula=['"✅ Pass"'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        "E2:E100", CellIsRule(operator="equal", formula=['"❌ Fail"'], fill=red_fill)
    )

    # Add timestamp formula: auto-fills when Status is updated
    for row in range(2, len(test_cases) + 2):
        ws[f"F{row}"].value = f'=IF(E{row}<>"", TEXT(NOW(), "yyyy-mm-dd hh:mm:ss"), "")'

    # --- Summary Sheet ---
    summary = wb.create_sheet(title="Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "Value"

    summary["A2"] = "Total Tests"
    summary["B2"] = f"=COUNTA('Gatekeeper Test Cases'!A2:A100)"

    summary["A3"] = "Passed"
    summary["B3"] = f"=COUNTIF('Gatekeeper Test Cases'!E2:E100,\"✅ Pass\")"

    summary["A4"] = "Failed"
    summary["B4"] = f"=COUNTIF('Gatekeeper Test Cases'!E2:E100,\"❌ Fail\")"

    summary["A5"] = "Success %"
    summary["B5"] = f"=IF(B2>0, B3/B2, 0)"

    # --- Pie Chart (Pass vs Fail) ---
    pie = PieChart()
    labels = Reference(summary, min_col=1, min_row=3, max_row=4)  # Passed/Failed
    data = Reference(summary, min_col=2, min_row=3, max_row=4)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Pass vs Fail"
    pie.dataLabels = True  # ✅ Show percentage labels
    summary.add_chart(pie, "D2")

    # --- Bar Chart (Pass/Fail per Feature) ---
    bar = BarChart()
    bar.title = "Pass/Fail per Feature"
    bar.x_axis.title = "Feature"
    bar.y_axis.title = "Count"

    labels = Reference(ws, min_col=1, min_row=2, max_row=len(test_cases) + 1)
    data = Reference(ws, min_col=5, min_row=2, max_row=len(test_cases) + 1)
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(labels)

    summary.add_chart(bar, "D20")

    # Save file
    wb.save(filename)
    print(
        f"✅ Test cases exported to {filename} with conditional formatting, timestamp, tester initials, summary sheet, pie chart (with % labels), and bar chart."
    )


if __name__ == "__main__":
    export_test_cases()
